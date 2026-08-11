"""Incident/escalation report generation: structured JSON payloads + styled
PDF files (reportlab) + Excel workbooks (openpyxl). Critical alerts get an
auto-filled escalation template at ingestion time.

Reports state what the system measured — sentiment split, concern scores, where
negative sentiment concentrated, and which posts drove it — and stop there. They
do not assert that a post is incitement or misinformation, because nothing in
the pipeline establishes that; a report an officer may attach to a case file is
the last place to blur the line between a measurement and a conclusion.

One payload, three renderings. `_build_payload` is the only thing that reads
the database or computes anything; `_render_pdf` and `_render_xlsx` reshape
what it returned and nothing else. That is what stops the document of record
and the working copy from disagreeing about the same window.
"""
import logging
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path
from statistics import mean
from xml.sax.saxutils import escape

from sqlmodel import select

from app.config import settings
from app.database import session_scope
from app.models import Alert, Post, Report
from app.services.serializers import iso, post_to_dict

from app.ml.score import band as _band

log = logging.getLogger("sentinel.reports")

# Recommended actions are keyed by the CONCERN BAND, not by a category. The
# system reports that a post is strongly negative and spreading; what an officer
# should do about that is a function of how severe and how widely read it is,
# which the score already captures. Keying these off an asserted category would
# have meant recommending a takedown request on the strength of a model's guess
# that a post "is incitement" — a claim it can no longer make.
RECOMMENDED_ACTIONS = {
    "critical": [
        "Review the post directly and confirm the reading before acting on it",
        "Notify the police station with jurisdiction over the referenced location",
        "Preserve evidence: archive post URL, screenshots and author profile",
        "Cross-check the author against prior incident records",
    ],
    "high": [
        "Add the author and associated hashtags to the active watchlist",
        "Monitor for escalation and for coordinated re-posting",
        "Brief community-liaison officers for the affected area",
    ],
    "elevated": [
        "Keep under passive monitoring; no action indicated on this post alone",
        "Track share velocity in case the sentiment spreads",
    ],
}


def escalation_template(post: Post) -> dict:
    """Pre-filled escalation packet attached to critical alerts automatically."""
    return {
        "incident_type": f"{post.sentiment_label} sentiment, concern {round(post.concern_score)}/100",
        "priority": "P1 — IMMEDIATE",
        "generated_by": "SENTINEL automated escalation",
        "generated_at": iso(datetime.now(timezone.utc).replace(tzinfo=None)),
        "platform": post.platform,
        "source_url": post.url,
        "location": post.location,
        "language": post.language,
        "author": {
            "handle": post.author_handle, "name": post.author_name,
            "followers": post.author_followers,
            "account_age_days": post.author_account_age_days,
        },
        "evidence": {
            "original_text": post.text,
            "english_translation": post.translation,
            "concern_score": post.concern_score,
            "sentiment": f"{post.sentiment_label} ({post.sentiment_confidence:.0%} confidence)",
            "how_the_score_was_built": (post.sentiment_consensus or {}).get("score_breakdown", []),
            "evidence_sources": [e.get("source") for e
                                 in (post.sentiment_consensus or {}).get("evidence", [])],
            "flags": post.hate_flags or [],
            "matched_keywords": post.keywords or [],
        },
        "recommended_actions": RECOMMENDED_ACTIONS.get(_band(post.concern_score), [])[:3],
        "note": "Automated triage output — requires analyst verification before action.",
    }


def _build_payload(period_hours: int) -> dict:
    since = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=period_hours)
    with session_scope() as s:
        posts = s.exec(select(Post).where(Post.created_at >= since)).all()
        alerts = s.exec(select(Alert).where(Alert.created_at >= since)).all()

    from app.services.network_service import get_network
    from app.services.trend_service import get_trends

    network = get_network(period_hours)
    trends = get_trends(period_hours)

    negative = [p for p in posts if p.sentiment_label == "negative"]
    top = sorted(posts, key=lambda p: -p.concern_score)[:6]
    label_counts = Counter(p.sentiment_label for p in posts)
    flagged = [p for p in posts if p.concern_score >= settings.ALERT_THRESHOLD]

    summary = (
        f"In the last {period_hours}h SENTINEL processed {len(posts)} posts across "
        f"{len({p.platform for p in posts})} platforms. Sentiment split "
        f"{label_counts.get('negative', 0)} negative / "
        f"{label_counts.get('neutral', 0)} neutral / "
        f"{label_counts.get('positive', 0)} positive; {len(flagged)} posts scored at or "
        f"above the concern threshold of {settings.ALERT_THRESHOLD}. "
        f"{sum(1 for a in alerts if a.severity == 'critical')} critical alerts were raised and "
        f"{len(network['clusters'])} coordinated amplification cluster(s) detected."
    )

    actions: list[str] = []
    for b, _ in Counter(_band(p.concern_score) for p in negative).most_common():
        actions.extend(RECOMMENDED_ACTIONS.get(b, [])[:2])

    return {
        "summary": summary,
        "period_hours": period_hours,
        "generated_at": iso(datetime.now(timezone.utc).replace(tzinfo=None)),
        "totals": {
            "posts": len(posts),
            "negative_posts": len(negative),
            "flagged_posts": len(flagged),
            "alerts": len(alerts),
            "critical_alerts": sum(1 for a in alerts if a.severity == "critical"),
            "avg_concern_score": round(mean(p.concern_score for p in posts), 1) if posts else 0,
        },
        "sentiment_distribution": dict(label_counts),
        "language_distribution": dict(Counter(p.language for p in posts)),
        "platform_distribution": dict(Counter(p.platform for p in posts)),
        "top_concern": [post_to_dict(p, full=True) for p in top],
        "coordinated_clusters": network["clusters"],
        "trending_hashtags": trends["hashtags"][:8],
        "regions": trends["regions"][:8],
        "recommended_actions": actions[:6],
    }


# ── PDF fonts ───────────────────────────────────────────────────────────────
# reportlab's built-in base-14 fonts (Helvetica and friends) are WinAnsi-only:
# they contain no Gujarati or Devanagari glyphs whatsoever. Every such
# character therefore renders as .notdef — the solid black box — which is
# indefensible in a product whose entire subject is Gujarati and Hindi social
# media. So Unicode fonts are registered and used for everything.
#
# Vendored rather than downloaded on first use. Report generation is on the
# request path, and a deployment that may end up air-gapped is the wrong place
# to discover that fonts.google.com is unreachable. Noto is SIL OFL licensed,
# which permits redistribution — see OFL.txt beside the files.
FONT_DIR = Path(__file__).resolve().parent.parent / "assets" / "fonts"

#: (family, regular file, bold file). Order matters: the first entry is the
#: default for everything that is not script-specific.
_FONT_FILES = (
    ("NotoSans", "NotoSans-Regular.ttf", "NotoSans-Bold.ttf"),
    ("NotoSansGujarati", "NotoSansGujarati-Regular.ttf", "NotoSansGujarati-Bold.ttf"),
    ("NotoSansDevanagari", "NotoSansDevanagari-Regular.ttf", "NotoSansDevanagari-Bold.ttf"),
    # Emoji has no script range and is never *preferred* — it is reached only
    # by the coverage search, as the last font that might have the character.
    # It earns its ~2 MB: a fifth of the corpus (4,061 of 19,213 seeded posts)
    # contains at least one emoji, and every one of them was a black box for
    # exactly the same reason the Gujarati was. Monochrome, because reportlab
    # cannot draw the colour (CBDT/COLR) tables anyway.
    ("NotoEmoji", "NotoEmoji-Regular.ttf", "NotoEmoji-Regular.ttf"),
)

BASE_FONT = "NotoSans"
FALLBACK_FONT = "Helvetica"
FALLBACK_FONT_BOLD = "Helvetica-Bold"

#: Which family to *prefer* for a character, by Unicode block. Preference, not
#: assignment — the final choice is made against real cmap coverage below,
#: because a preferred font that lacks the character is no better than
#: Helvetica was.
_SCRIPT_RANGES = (
    ((0x0A80, 0x0AFF), "NotoSansGujarati"),   # Gujarati
    ((0x0900, 0x097F), "NotoSansDevanagari"),  # Devanagari (Hindi, Marathi)
    ((0xA8E0, 0xA8FF), "NotoSansDevanagari"),  # Devanagari Extended
)

#: Result of the one registration attempt: {"ok": bool, "coverage": {...}}.
#: Cached because registering parses six TTFs, and reports are generated on a
#: request.
_font_state: dict | None = None


def _register_fonts() -> dict:
    """Register the Noto families with reportlab, once per process.

    Never raises. A missing or corrupt font file downgrades the report to the
    old Helvetica behaviour and logs why — an English report with a font
    warning beats no report at all, and the Indic text was already unreadable
    in that case rather than newly broken.
    """
    global _font_state
    if _font_state is not None:
        return _font_state

    state: dict = {"ok": False, "coverage": {}}
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont, TTFontFile
    except ImportError:
        log.warning("reportlab missing — cannot register Unicode fonts")
        _font_state = state
        return state

    for family, regular, bold in _FONT_FILES:
        regular_path, bold_path = FONT_DIR / regular, FONT_DIR / bold
        try:
            pdfmetrics.registerFont(TTFont(family, str(regular_path)))
            pdfmetrics.registerFont(TTFont(f"{family}-Bold", str(bold_path)))
            # Without the family mapping, a <b> tag inside a run set to this
            # font silently falls back to Helvetica-Bold and the bold words in
            # a Gujarati heading go back to being black boxes.
            pdfmetrics.registerFontFamily(
                family, normal=family, bold=f"{family}-Bold",
                italic=family, boldItalic=f"{family}-Bold")
            # The cmap is what decides which font renders a character. Read it
            # from the file rather than assuming the obvious: Noto's Indic
            # fonts carry barely any Latin (Gujarati has ~116 codepoints and no
            # ASCII letters), so "use the Gujarati font for a Gujarati post"
            # would tofu every English word in a code-mixed one.
            state["coverage"][family] = set(TTFontFile(str(regular_path)).charToGlyph)
        except Exception:
            log.warning("could not register PDF font %s from %s — "
                        "falling back to %s for that family",
                        family, regular_path, FALLBACK_FONT, exc_info=True)

    state["ok"] = BASE_FONT in state["coverage"]
    if not state["ok"]:
        log.warning("Unicode PDF fonts unavailable (looked in %s) — reports will "
                    "render with %s, and non-Latin text will not display "
                    "correctly", FONT_DIR, FALLBACK_FONT)
    _font_state = state
    return state


def _font_for_char(char: str, coverage: dict) -> str:
    """The registered font that can actually draw this character."""
    code = ord(char)
    preferred = next((family for (low, high), family in _SCRIPT_RANGES
                      if low <= code <= high), None)
    if preferred and code in coverage.get(preferred, ()):
        return preferred
    if code in coverage.get(BASE_FONT, ()):
        return BASE_FONT
    # Neither the script font nor the Latin one has it — anything that does is
    # better than a black box.
    for family, covered in coverage.items():
        if code in covered:
            return family
    return BASE_FONT


def _markup(text: str, coverage: dict | None) -> str:
    """User text → safe reportlab paragraph markup.

    Does the two things that were separately broken:

    1. **Escapes it.** `Paragraph` parses its content as mini-XML, so a handle
       or post body containing a bare "<" raised
       `ValueError: paraparser: syntax error ... unclosed tags` and took the
       whole report down. That failure had nothing to do with Unicode; it just
       happened to be reported alongside it.

    2. **Splits it into per-font runs.** Splitting per *character block* rather
       than per paragraph is what makes code-mixed text work: "Surat માં
       tension" needs the Latin font for two words and the Gujarati font for
       one, and no single Noto file covers both.

    With no Unicode fonts registered this degrades to escaping alone, which is
    exactly the old rendering minus the crash.
    """
    if not text:
        return ""
    if not coverage:
        return escape(text)

    runs: list[tuple[str, list[str]]] = []
    for char in text:
        font = _font_for_char(char, coverage)
        if runs and runs[-1][0] == font:
            runs[-1][1].append(char)
        else:
            runs.append((font, [char]))
    return "".join(
        f'<font name="{font}">{escape("".join(chars))}</font>' for font, chars in runs
    )


def _render_pdf(report: Report) -> str:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        log.warning("reportlab missing — skipping PDF render")
        return ""

    path = settings.REPORTS_DIR / f"{report.id}.pdf"
    p = report.payload

    fonts = _register_fonts()
    coverage = fonts["coverage"] if fonts["ok"] else None
    base_font = BASE_FONT if fonts["ok"] else FALLBACK_FONT
    bold_font = f"{BASE_FONT}-Bold" if fonts["ok"] else FALLBACK_FONT_BOLD

    def mk(text) -> str:
        """Every piece of user-controlled text in this document goes through
        here. Static English chrome does not need to, but is harmless."""
        return _markup("" if text is None else str(text), coverage)

    styles = getSampleStyleSheet()
    # fontName is set explicitly on each style rather than inherited: the
    # sample stylesheet's parents are all Helvetica, so anything left to
    # inherit would quietly keep rendering Indic text as black boxes.
    h1 = ParagraphStyle("h1", parent=styles["Title"], textColor=colors.HexColor("#0F1420"),
                        fontSize=18, fontName=bold_font)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#14B8C4"),
                        fontName=bold_font)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontName=base_font)

    doc = SimpleDocTemplate(str(path), pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    flow = [
        Paragraph("SENTINEL — Public Sentiment Report", h1),
        # The title is operator-supplied and the summary can quote a post, so
        # both are user text as far as this document is concerned.
        Paragraph(f"{mk(report.title)} • generated {mk(p.get('generated_at', ''))} "
                  f"• window {mk(p.get('period_hours'))}h", body),
        Spacer(1, 6 * mm),
        Paragraph("Executive Summary", h2),
        Paragraph(mk(p.get("summary", "")), body),
        Spacer(1, 4 * mm),
        Paragraph("Sentiment Breakdown", h2),
    ]
    dist = p.get("sentiment_distribution", {})
    # Table cells are not parsed as markup, so they cannot carry per-run fonts.
    # Sentiment labels are a fixed English vocabulary, so one font for the whole
    # table is correct here — and it is the Unicode one, so the table matches
    # the rest of the document.
    table = Table([["Sentiment", "Posts"]] + [[k, str(v)] for k, v in dist.items()], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F1420")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#94A3B8")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTNAME", (0, 1), (-1, -1), base_font),
    ]))
    flow.append(table)

    flow.append(Spacer(1, 4 * mm))
    flow.append(Paragraph("Highest-Concern Posts", h2))
    for t in p.get("top_concern", [])[:5]:
        # Handle, location and label are all user- or platform-supplied and any
        # of them can be non-Latin — a Gujarati place name in `location` was
        # tofu even when the post body happened to be English.
        headline = "[{}] {}".format(t["concern_score"], t["sentiment_label"])
        flow.append(Paragraph(
            f"<b>{mk(headline)}</b> — {mk(t['platform'])} @{mk(t['author_handle'])} "
            f"({mk(t['language'])}, {mk(t['location'] or 'n/a')})", body))
        flow.append(Paragraph(mk(t.get("translation") or t.get("text", "")), body))
        flow.append(Spacer(1, 2 * mm))

    clusters = p.get("coordinated_clusters", [])
    if clusters:
        flow.append(Paragraph("Coordinated Amplification", h2))
        for c in clusters[:4]:
            flow.append(Paragraph(
                f"<b>{mk(c['id'])} — {mk(c['label'])}</b> (confidence {c['confidence']:.0%}, "
                f"{len(c['accounts'])} accounts): {mk('; '.join(c['why']))}", body))
            flow.append(Spacer(1, 2 * mm))

    flow.append(Paragraph("Recommended Actions", h2))
    for a in p.get("recommended_actions", []):
        flow.append(Paragraph(f"• {mk(a)}", body))

    doc.build(flow)
    return str(path)


def _render_xlsx(report: Report) -> str:
    """The same payload as the PDF, as a workbook analysts can actually work in.

    The PDF is the document of record and stays exactly as it was. This is the
    other half of the same data: a PDF cannot be sorted, filtered or pasted
    into a case file, and the highest-concern list is precisely the table an
    analyst wants to re-rank. So the sheet that matters carries every scored
    post rather than the PDF's first five, and ships with the filter, frozen
    header and score gradient already applied — a workbook that needs three
    manual steps before it is readable does not get used.

    Nothing is recomputed. Every value here is reshaped from `report.payload`,
    which `_build_payload` already produced for the JSON and the PDF, so the
    three renderings cannot disagree about what happened in the window.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.worksheet import Worksheet
    except ImportError:
        log.warning("openpyxl missing — skipping XLSX render")
        return ""

    path = settings.REPORTS_DIR / f"{report.id}.xlsx"
    p = report.payload

    # The PDF's palette, so the two renderings of one report look related.
    header_fill = PatternFill("solid", fgColor="0F1420")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    def head(ws: "Worksheet", row: int, labels: list[str], col: int = 1) -> None:
        """Write one styled header row. Every table on every sheet goes through
        here, so a sheet added later cannot quietly look different."""
        for offset, label in enumerate(labels):
            cell = ws.cell(row=row, column=col + offset, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def autosize(ws: "Worksheet", limit: int = 60) -> None:
        """Width from the longest cell in each column, capped.

        Uncapped, one 400-character post body makes a column wider than the
        screen and the sheet is unreadable in a different way than before.
        """
        widest: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                longest = max(len(line) for line in str(cell.value).split("\n"))
                if longest > widest.get(cell.column, 0):
                    widest[cell.column] = longest
        for column, width in widest.items():
            ws.column_dimensions[get_column_letter(column)].width = min(width + 2, limit)

    totals = p.get("totals", {})
    wb = Workbook()

    # ── Summary ─────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    head(ws, 1, ["Metric", "Value"])
    for label, value in [
        ("Report Title", report.title),
        ("Report Kind", report.kind),
        ("Window (hours)", p.get("period_hours", "")),
        ("Generated At", p.get("generated_at", "")),
        ("Posts Processed", totals.get("posts", 0)),
        ("Negative Posts", totals.get("negative_posts", 0)),
        ("Flagged Posts", totals.get("flagged_posts", 0)),
        ("Alerts", totals.get("alerts", 0)),
        ("Critical Alerts", totals.get("critical_alerts", 0)),
        ("Coordinated Clusters", len(p.get("coordinated_clusters", []))),
        ("Average Concern Score", totals.get("avg_concern_score", 0)),
    ]:
        ws.append([label, value])
    ws.append([])
    ws.append(["Executive Summary", p.get("summary", "")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    # The summary is a paragraph, so it wraps rather than widening the sheet
    # to the length of a sentence.
    ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)
    ws.column_dimensions["B"].width = 90

    # ── Sentiment Breakdown ─────────────────────────────────────────────────
    ws = wb.create_sheet("Sentiment Breakdown")
    head(ws, 1, ["Sentiment", "Count", "%"])
    distribution = p.get("sentiment_distribution", {})
    total_classified = sum(distribution.values())
    for label, count in distribution.items():
        share = round(100 * count / total_classified, 1) if total_classified else 0
        ws.append([label, count, share])
    autosize(ws)

    # ── Highest-Concern Posts ───────────────────────────────────────────────
    ws = wb.create_sheet("Highest-Concern Posts")
    concern_columns = ["Concern Score", "Sentiment", "Confidence", "Platform",
                       "Author Handle", "Language", "Location", "Post Text",
                       "Translation", "URL", "Matched Keywords", "Created At"]
    head(ws, 1, concern_columns)
    for post in p.get("top_concern", []):
        ws.append([
            post.get("concern_score", 0),
            post.get("sentiment_label", ""),
            post.get("sentiment_confidence", 0),
            post.get("platform", ""),
            post.get("author_handle", ""),
            post.get("language", ""),
            post.get("location") or "",
            post.get("text", ""),
            post.get("translation") or "",
            post.get("url", ""),
            ", ".join(post.get("keywords") or []),
            post.get("created_at", ""),
        ])
    last_row = ws.max_row
    last_column = get_column_letter(len(concern_columns))
    ws.auto_filter.ref = f"A1:{last_column}{last_row}"
    # Below the header, so the columns stay labelled while scrolling a long
    # list — which is the whole reason this sheet is not capped at five.
    ws.freeze_panes = "A2"
    if last_row > 1:
        # Red at the top of the range, green at the bottom: the eye should land
        # on the worst row without reading a number.
        ws.conditional_formatting.add(
            f"A2:A{last_row}",
            ColorScaleRule(start_type="min", start_color="63BE7B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="F8696B"),
        )
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws, limit=50)

    # ── Platform & Language ─────────────────────────────────────────────────
    # Two independent distributions, side by side with a spacer column: both
    # are short, and a sheet each would be three clicks to compare.
    ws = wb.create_sheet("Platform & Language")
    head(ws, 1, ["Platform", "Posts"], col=1)
    head(ws, 1, ["Language", "Posts"], col=4)
    platforms = list(p.get("platform_distribution", {}).items())
    languages = list(p.get("language_distribution", {}).items())
    for index in range(max(len(platforms), len(languages))):
        row = index + 2
        if index < len(platforms):
            ws.cell(row=row, column=1, value=platforms[index][0])
            ws.cell(row=row, column=2, value=platforms[index][1])
        if index < len(languages):
            ws.cell(row=row, column=4, value=languages[index][0])
            ws.cell(row=row, column=5, value=languages[index][1])
    autosize(ws)

    # ── Trending & Regions ──────────────────────────────────────────────────
    ws = wb.create_sheet("Trending & Regions")
    head(ws, 1, ["Hashtag / Term", "Mentions", "Change %", "Spiking", "Top Sentiment"], col=1)
    head(ws, 1, ["Region", "Posts", "Negative", "Avg Concern Score"], col=7)
    hashtags = p.get("trending_hashtags", [])
    regions = p.get("regions", [])
    for index in range(max(len(hashtags), len(regions))):
        row = index + 2
        if index < len(hashtags):
            tag = hashtags[index]
            ws.cell(row=row, column=1, value=tag.get("term", ""))
            ws.cell(row=row, column=2, value=tag.get("count", 0))
            ws.cell(row=row, column=3, value=tag.get("change_pct", 0))
            ws.cell(row=row, column=4, value="yes" if tag.get("spiking") else "no")
            ws.cell(row=row, column=5, value=tag.get("top_label", ""))
        if index < len(regions):
            region = regions[index]
            ws.cell(row=row, column=7, value=region.get("name", ""))
            ws.cell(row=row, column=8, value=region.get("count", 0))
            # `threats` is the trend service's long-standing key for "posts in
            # this region that came out negative" — the column is named for
            # what it counts rather than for what the key is called.
            ws.cell(row=row, column=9, value=region.get("threats", 0))
            ws.cell(row=row, column=10, value=region.get("avg_concern", 0))
    autosize(ws)

    # ── Coordinated Clusters ────────────────────────────────────────────────
    # Omitted entirely when nothing was detected. An empty sheet headed
    # "Coordinated Clusters" reads as a rendering failure rather than as a
    # quiet window.
    clusters = p.get("coordinated_clusters", [])
    if clusters:
        ws = wb.create_sheet("Coordinated Clusters")
        head(ws, 1, ["Cluster ID", "Label", "Confidence", "Accounts", "Why Flagged"])
        for cluster in clusters:
            ws.append([
                cluster.get("id", ""),
                cluster.get("label", ""),
                cluster.get("confidence", 0),
                len(cluster.get("accounts") or []),
                "; ".join(cluster.get("why") or []),
            ])
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        autosize(ws, limit=70)

    # ── Recommended Actions ─────────────────────────────────────────────────
    ws = wb.create_sheet("Recommended Actions")
    head(ws, 1, ["#", "Recommended Action", "Status"])
    for index, action in enumerate(p.get("recommended_actions", []), start=1):
        ws.append([index, action, ""])
    # Deliberately blank: the workbook is handed round a shift and the status
    # column is where it gets filled in. It exists so nobody adds one in a
    # different place on every copy.
    autosize(ws, limit=80)

    wb.save(str(path))
    return str(path)


def generate_report(title: str = "", period_hours: int = 24, kind: str = "incident") -> Report:
    payload = _build_payload(period_hours)
    report = Report(
        title=title or f"Situation Report — last {period_hours}h",
        kind=kind, period_hours=period_hours, payload=payload,
    )
    with session_scope() as s:
        s.add(report)
        s.commit()
        s.refresh(report)
        report.pdf_path = _render_pdf(report)
        report.xlsx_path = _render_xlsx(report)
        s.add(report)
        s.commit()
        s.refresh(report)
        # detach a plain copy
        s.expunge(report)
    return report
